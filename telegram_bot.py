import os
import logging
import openpyxl
from datetime import datetime, timedelta
from telegram import Update
from telegram.ext import Application, MessageHandler, CommandHandler, filters, ContextTypes
from tao_bao_cao import tao_bao_cao, doc_du_lieu
from tao_bao_cao_doanhthu import tong_hop_doanh_thu, doc_file_don_le_doanhthu, them_vao_doanh_thu
from doc_file_don_le import doc_file_don_le, them_vao_tong_hop
import pandas as pd

try:
    from google_drive import upload_file, download_file, list_files_on_drive
    USE_DRIVE = bool(os.environ.get("GOOGLE_CREDENTIALS"))
except Exception:
    USE_DRIVE = False

BOT_TOKEN = "8816246373:AAFNJr6KWKxXWgQwBYQu5nFqGOl574m-q-c"
CHAT_ID = None
DOWNLOAD_DIR = "downloads"
MASTER_FILENAME = "tong_hop_san_luong.xlsx"
MASTER_FILE = os.path.join(DOWNLOAD_DIR, MASTER_FILENAME)
DOANHTHU_FILENAME = "master_doanh_thu.xlsx"
DOANHTHU_FILE = os.path.join(DOWNLOAD_DIR, DOANHTHU_FILENAME)

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(message)s")
os.makedirs(DOWNLOAD_DIR, exist_ok=True)


def sync_master_from_drive():
    if not USE_DRIVE:
        return
    try:
        result = download_file(MASTER_FILENAME, MASTER_FILE)
        if result:
            logging.info("Synced master file from Drive")
    except Exception as e:
        logging.warning(f"Khong the tai master file tu Drive: {e}")


def sync_master_to_drive():
    if not USE_DRIVE:
        return
    try:
        if os.path.exists(MASTER_FILE):
            upload_file(MASTER_FILE, MASTER_FILENAME)
            logging.info("Synced master file to Drive")
    except Exception as e:
        logging.warning(f"Khong the upload master file len Drive: {e}")


def la_file_don_le(filename):
    name = filename.lower().strip()
    return name.startswith("in ") or name.startswith("in_")


async def luu_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.message
    if not msg or not msg.document:
        return
    filename = msg.document.file_name or ""
    if not (filename.endswith(".xlsx") or filename.endswith(".xls")):
        return
    if CHAT_ID and msg.chat_id != CHAT_ID:
        return
    try:
        file = await context.bot.get_file(msg.document.file_id)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        save_name = f"{ts}_{filename}"
        input_path = os.path.join(DOWNLOAD_DIR, save_name)
        await file.download_to_drive(input_path)
        logging.info(f"Da luu: {save_name}")
        if USE_DRIVE:
            try:
                upload_file(input_path, save_name)
            except Exception as e:
                logging.warning(f"Khong upload duoc file goc: {e}")
        if la_file_don_le(filename):
            ngay_gui = datetime.now()
            ten_sale = msg.from_user.full_name if msg.from_user else ""
            # Tong hop san luong
            if not os.path.exists(MASTER_FILE):
                sync_master_from_drive()
            du_lieu, ten_kh = doc_file_don_le(input_path, ten_sale)
            tong = them_vao_tong_hop(du_lieu, MASTER_FILE)
            sync_master_to_drive()
            # Tong hop doanh thu
            try:
                if not os.path.exists(DOANHTHU_FILE):
                    try:
                        download_file(DOANHTHU_FILENAME, DOANHTHU_FILE)
                    except Exception:
                        pass
                dong_dt = doc_file_don_le_doanhthu(input_path, ten_sale)
                dong_dt["Ngày gửi đơn"] = ngay_gui
                them_vao_doanh_thu(dong_dt, DOANHTHU_FILE)
                if USE_DRIVE:
                    upload_file(DOANHTHU_FILE, DOANHTHU_FILENAME)
            except Exception as e:
                logging.warning(f"Khong them duoc vao doanh thu: {e}")
            await msg.reply_text(
                f"Da luu va tong hop: {filename}\n"
                f"Khach: {ten_kh} | {len(du_lieu)} dong | Tong: {tong} dong"
            )
        else:
            await msg.reply_text(f"Da luu file: {filename}")
    except Exception as e:
        logging.error(f"Loi luu file: {e}")
        await msg.reply_text(f"Loi khi xu ly file: {str(e)}")


def tinh_khoang_ngay(args):
    now = datetime.now()
    if not args:
        today = now.date()
        days_since_monday = today.weekday()
        last_monday = today - timedelta(days=days_since_monday + 7)
        last_sunday = last_monday + timedelta(days=6)
        tu_ngay = datetime.combine(last_monday, datetime.min.time())
        den_ngay = datetime.combine(last_sunday, datetime.max.time().replace(microsecond=0))
        ten_ky = f"Tuan {last_monday.strftime('%d/%m')} - {last_sunday.strftime('%d/%m/%Y')}"
        return tu_ngay, den_ngay, ten_ky
    if args[0].lower() == "thang":
        first_this_month = now.replace(day=1)
        last_month_end = first_this_month - timedelta(days=1)
        last_month_start = last_month_end.replace(day=1)
        tu_ngay = datetime.combine(last_month_start, datetime.min.time())
        den_ngay = datetime.combine(last_month_end, datetime.max.time().replace(microsecond=0))
        ten_ky = f"Thang {last_month_start.strftime('%m/%Y')}"
        return tu_ngay, den_ngay, ten_ky
    if len(args) >= 2:
        try:
            year = now.year
            d1 = datetime.strptime(f"{args[0]}/{year}", "%d/%m/%Y")
            d2 = datetime.strptime(f"{args[1]}/{year}", "%d/%m/%Y")
            tu_ngay = d1.replace(hour=0, minute=0, second=0)
            den_ngay = d2.replace(hour=23, minute=59, second=59)
            ten_ky = f"{d1.strftime('%d/%m')} - {d2.strftime('%d/%m/%Y')}"
            return tu_ngay, den_ngay, ten_ky
        except ValueError:
            return None, None, None
    return None, None, None


async def tao_bao_cao_tuan(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.message
    if CHAT_ID and msg.chat_id != CHAT_ID:
        return
    args = context.args or []
    tu_ngay, den_ngay, ten_ky = tinh_khoang_ngay(args)
    if tu_ngay is None:
        await msg.reply_text("Cu phap:\n/baocao - tuan truoc\n/baocao thang - thang truoc\n/baocao 18/5 24/5 - khoang ngay")
        return
    await msg.reply_text(f"Dang tong hop san luong: {ten_ky}...")
    if not os.path.exists(MASTER_FILE):
        sync_master_from_drive()
    if not os.path.exists(MASTER_FILE):
        await msg.reply_text(f"Khong tim thay du lieu trong: {ten_ky}")
        return
    try:
        try:
            df = pd.read_excel(MASTER_FILE, sheet_name="Tổng hợp sản lượng")
        except Exception:
            df = pd.read_excel(MASTER_FILE, sheet_name=0)
        if "Ngay dua don" in df.columns:
            df = df.rename(columns={"Ngay dua don": "Ngày đưa đơn"})
        df["Ngày đưa đơn"] = pd.to_datetime(df["Ngày đưa đơn"], errors="coerce")
        df_filtered = df[(df["Ngày đưa đơn"] >= tu_ngay) & (df["Ngày đưa đơn"] <= den_ngay)]
        if df_filtered.empty:
            await msg.reply_text(f"Khong co du lieu trong: {ten_ky}")
            return
        tmp = os.path.join(DOWNLOAD_DIR, "temp_filtered.xlsx")
        df_filtered.to_excel(tmp, sheet_name="Tổng hợp sản lượng", index=False)
        ten_file = ten_ky.replace("/", "-").replace(" ", "_")
        output_path = os.path.join(DOWNLOAD_DIR, f"BaoCao_{ten_file}.xlsx")
        tao_bao_cao(tmp, output_path)
        with open(output_path, "rb") as f:
            await msg.reply_document(
                document=f,
                filename=os.path.basename(output_path),
                caption=f"Bao cao san luong {ten_ky} ({len(df_filtered)} dong)"
            )
    except Exception as e:
        logging.error(f"Loi tao bao cao: {e}")
        await msg.reply_text(f"Loi: {str(e)}")


async def tao_bao_cao_doanh_thu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.message
    if CHAT_ID and msg.chat_id != CHAT_ID:
        return
    args = context.args or []
    if not args:
        await msg.reply_text("Cu phap:\n/doanhthu 5 - thang 5\n/doanhthu 4 2026 - thang 4/2026")
        return
    try:
        thang = int(args[0])
        nam = int(args[1]) if len(args) >= 2 else datetime.now().year
    except ValueError:
        await msg.reply_text("Thang phai la so. Vi du: /doanhthu 4")
        return
    if not (1 <= thang <= 12):
        await msg.reply_text("Thang phai tu 1 den 12.")
        return
    tu_ngay = datetime(nam, thang, 1)
    den_ngay = datetime(nam, thang + 1, 1) - timedelta(seconds=1) if thang < 12 else datetime(nam + 1, 1, 1) - timedelta(seconds=1)
    ten_thang = f"Thang {thang}/{nam}"
    # Tai master doanh thu
    if not os.path.exists(DOANHTHU_FILE):
        try:
            download_file(DOANHTHU_FILENAME, DOANHTHU_FILE)
        except Exception:
            pass
    if not os.path.exists(DOANHTHU_FILE):
        await msg.reply_text("Chua co du lieu doanh thu. Hay gui file don hang vao nhom truoc.")
        return
    await msg.reply_text(f"Dang tong hop doanh thu {ten_thang}...")
    try:
        df = pd.read_excel(DOANHTHU_FILE, sheet_name="Quản lý")
        col_ngay = None
        for c in df.columns:
            if "ngay" in str(c).lower() and "gui" in str(c).lower():
                col_ngay = c
                break
        if not col_ngay:
            for c in df.columns:
                if "in" in str(c).lower() and "don" in str(c).lower():
                    col_ngay = c
                    break
        if col_ngay:
            df[col_ngay] = pd.to_datetime(df[col_ngay], errors="coerce")
            df_filter = df[(df[col_ngay] >= tu_ngay) & (df[col_ngay] <= den_ngay)]
        else:
            df_filter = df
        if df_filter.empty:
            await msg.reply_text(f"Khong co du lieu doanh thu {ten_thang}.")
            return
        output_path = os.path.join(DOWNLOAD_DIR, f"DoanhThu_Thang{thang:02d}_{nam}.xlsx")
        import openpyxl as _opx
        from tao_bao_cao_doanhthu import viet_sheet_quan_ly, viet_sheet_cong_no, viet_sheet_dthu
        wb = _opx.Workbook()
        wb.remove(wb.active)
        col_ten = "Tên khách "
        col_amount = "Amount"
        col_sale = "Sale"
        col_indon = "In đơn"
        col_inv = "INV Date"
        col_ngay = "Ngày gửi đơn"
        col_bit = "Số Bit "
        viet_sheet_quan_ly(wb, df_filter, col_ten, col_amount, col_sale, col_indon, col_inv, col_ngay, col_bit)
        if col_ten in df_filter.columns and col_amount in df_filter.columns:
            viet_sheet_cong_no(wb, df_filter, col_ten, col_amount)
        if col_sale in df_filter.columns and col_amount in df_filter.columns:
            viet_sheet_dthu(wb, df_filter, col_sale, col_amount)
        wb.save(output_path)
        tong_amount = df_filter["Amount"].sum() if "Amount" in df_filter.columns else 0
        with open(output_path, "rb") as f:
            await msg.reply_document(
                document=f,
                filename=os.path.basename(output_path),
                caption=f"Bao cao doanh thu {ten_thang}\nSo don: {len(df_filter)} | Tong amount: {tong_amount:,.0f}"
            )
    except Exception as e:
        logging.error(f"Loi tao bao cao doanh thu: {e}")
        await msg.reply_text(f"Loi: {str(e)}")


async def xem_tonghop(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = update.message
    if CHAT_ID and msg.chat_id != CHAT_ID:
        return
    args = context.args or []
    now = datetime.now()
    if not os.path.exists(MASTER_FILE):
        sync_master_from_drive()
    if not os.path.exists(MASTER_FILE):
        await msg.reply_text("Chua co du lieu tong hop. Hay gui file don hang vao nhom truoc.")
        return
    try:
        try:
            df = pd.read_excel(MASTER_FILE, sheet_name="Tổng hợp sản lượng")
        except Exception:
            df = pd.read_excel(MASTER_FILE, sheet_name=0)
        if "Ngay dua don" in df.columns:
            df = df.rename(columns={"Ngay dua don": "Ngày đưa đơn"})
        df["Ngày đưa đơn"] = pd.to_datetime(df["Ngày đưa đơn"], errors="coerce")
        if not args:
            ngay = now.date()
            df_filter = df[df["Ngày đưa đơn"].dt.date == ngay]
            ten_ky = f"ngay {ngay.strftime('%d/%m/%Y')}"
        elif args[0] == "hom nay":
            ngay = now.date()
            df_filter = df[df["Ngày đưa đơn"].dt.date == ngay]
            ten_ky = f"ngay {ngay.strftime('%d/%m/%Y')}"
        elif args[0] == "tuan":
            today = now.date()
            monday = today - timedelta(days=today.weekday())
            df_filter = df[df["Ngày đưa đơn"].dt.date >= monday]
            ten_ky = f"tuan nay tu {monday.strftime('%d/%m')}"
        elif len(args) >= 2:
            try:
                year = now.year
                d1 = datetime.strptime(f"{args[0]}/{year}", "%d/%m/%Y").date()
                d2 = datetime.strptime(f"{args[1]}/{year}", "%d/%m/%Y").date()
                df_filter = df[(df["Ngày đưa đơn"].dt.date >= d1) & (df["Ngày đưa đơn"].dt.date <= d2)]
                ten_ky = f"{d1.strftime('%d/%m')} - {d2.strftime('%d/%m/%Y')}"
            except ValueError:
                df_filter = df
                ten_ky = "tat ca"
        else:
            try:
                year = now.year
                d = datetime.strptime(f"{args[0]}/{year}", "%d/%m/%Y").date()
                df_filter = df[df["Ngày đưa đơn"].dt.date == d]
                ten_ky = f"ngay {d.strftime('%d/%m/%Y')}"
            except ValueError:
                df_filter = df
                ten_ky = "tat ca"
        if df_filter.empty:
            await msg.reply_text(f"Khong co du lieu {ten_ky}.")
            return
        tmp = os.path.join(DOWNLOAD_DIR, "temp_tonghop.xlsx")
        df_filter.to_excel(tmp, sheet_name="Tổng hợp sản lượng", index=False)
        output_path = os.path.join(DOWNLOAD_DIR, f"TongHop_{now.strftime('%d%m%Y')}.xlsx")
        tao_bao_cao(tmp, output_path)
        # Them sheet chi tiet vao dau
        wb = openpyxl.load_workbook(output_path)
        ws_raw = wb.create_sheet("Chi tiet don hang", 0)
        cols = list(df_filter.columns)
        ws_raw.append(cols)
        for _, row in df_filter.iterrows():
            ws_raw.append([row[c] for c in cols])
        wb.save(output_path)
        with open(output_path, "rb") as f:
            await msg.reply_document(
                document=f,
                filename=os.path.basename(output_path),
                caption=f"Tong hop san luong {ten_ky} ({len(df_filter)} dong)"
            )
    except Exception as e:
        logging.error(f"Loi xem tonghop: {e}")
        await msg.reply_text(f"Loi: {str(e)}")


def main():
    if USE_DRIVE:
        logging.info("Google Drive: ON - syncing master files...")
        sync_master_from_drive()
        try:
            download_file(DOANHTHU_FILENAME, DOANHTHU_FILE)
        except Exception:
            pass
    else:
        logging.info("Google Drive: OFF")
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(MessageHandler(filters.Document.ALL, luu_file))
    app.add_handler(CommandHandler("baocao", tao_bao_cao_tuan))
    app.add_handler(CommandHandler("doanhthu", tao_bao_cao_doanh_thu))
    app.add_handler(CommandHandler("tonghop", xem_tonghop))
    print("Bot dang chay...")
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
