import sys
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def style_header(cell, bg="1F4E79", fg="FFFFFF"):
    cell.font = Font(name="Arial", bold=True, color=fg, size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_total(cell):
    cell.font = Font(name="Arial", bold=True, size=10)
    cell.fill = PatternFill("solid", start_color="D6E4F0")
    thin = Side(style="thin")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _parse_date_str(s):
    """Chuyen chuoi ngay (30-5, 30.5, 30/5) thanh datetime."""
    if not s:
        return None
    s = str(s).strip()
    for sep in ["-", ".", "/"]:
        if sep in s:
            parts = s.split(sep)
            if len(parts) >= 2:
                try:
                    day, month = int(parts[0]), int(parts[1])
                    year = int(parts[2]) if len(parts) > 2 else datetime.now().year
                    return datetime(year, month, day)
                except Exception:
                    pass
    return None


def doc_file_don_le_doanhthu(filepath, ten_sale=None):
    """Doc file don hang le de lay thong tin doanh thu."""
    from doc_file_don_le import tim_ten_khach, tim_header_row, tim_ngay, parse_ngay
    import re

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    rows = [r for r in all_rows if any(v is not None for v in r)]
    if len(rows) < 2:
        raise ValueError("File qua ngan")

    header_idx = tim_header_row(rows) or 3

    # Doc ten sale tu file (o ben phai chu "Sale")
    for _row in rows[:4]:
        for _j, _v in enumerate(_row):
            if _v and str(_v).lower().strip() == 'sale':
                if _j + 1 < len(_row) and _row[_j + 1]:
                    ten_sale = str(_row[_j + 1]).strip()
                break

    # Tim ten khach dung ham chung
    ten_kh = tim_ten_khach(rows, header_idx)

    # Tim ngay
    ngay_in_don, inv_date = tim_ngay(rows, header_idx)

    # Tim Amount: label "Amount" -> gia tri hang tiep theo
    amount = None
    for i, row in enumerate(rows[:header_idx]):
        for j, v in enumerate(row):
            if v and str(v).strip().lower() == "amount":
                if i + 1 < len(rows) and j < len(rows[i+1]):
                    val = rows[i+1][j]
                    if val is not None:
                        try:
                            amount = float(val)
                        except Exception:
                            pass
                break
        if amount is not None:
            break

    # Format Vera: dong tong o cuoi
    if amount is None:
        for row in reversed(rows):
            for v in row:
                if v and str(v).strip().lower().replace("\u1ed5", "o").replace("\u00f4", "o") in ["tong", "total", "t\u1ed5ng"]:
                    for val in reversed(row):
                        if val is not None:
                            try:
                                amount = float(val)
                                break
                            except Exception:
                                pass
                    break
            if amount is not None:
                break

    return {
        "Tên khách ": ten_kh,
        "In đơn": ngay_in_don,
        "INV Date": inv_date,
        "Ngày gửi đơn": ngay_in_don,
        "Amount": amount or 0,
        "Số Bit ": None,
        "Sale": ten_sale or "",
    }


def doc_file_doanhthu(filepath, ten_sale=None):
    """
    Doc file doanh thu - ho tro 2 loai:
    - File tong hop (co sheet 'Quan ly'): tra ve DataFrame
    - File don le (in xxx.xlsx): tra ve 1 dong tu dong trich xuat
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # Kiem tra co sheet Quan ly khong
    has_quan_ly = any("quan" in s.lower() or ("ly" in s.lower() and len(s) < 15)
                      for s in wb.sheetnames)

    if has_quan_ly:
        # File tong hop cu
        sheet_name = None
        for s in wb.sheetnames:
            if "quan" in s.lower() or "ly" in s.lower():
                sheet_name = s
                break
        if not sheet_name:
            sheet_name = wb.sheetnames[0]

        df = pd.read_excel(filepath, sheet_name=sheet_name, header=0)
        df.columns = [str(c).strip() for c in df.columns]

        col_ten = col_amount = col_sale = col_indon = col_inv = col_ngay = col_bit = None
        for col in df.columns:
            cl = col.lower().replace(" ", "").replace("\n", "")
            if "ten" in cl and ("khach" in cl or "kh" in cl):
                col_ten = col
            elif col.lower() == "amount":
                col_amount = col
            elif "sale" in cl:
                col_sale = col
            elif "indon" in cl or "inđon" in cl or ("in" in cl and "don" in cl):
                col_indon = col
            elif "inv" in cl:
                col_inv = col
            elif "ngay" in cl and "gui" in cl:
                col_ngay = col
            elif "bit" in cl:
                col_bit = col

        if not col_amount:
            raise ValueError(f"Khong tim thay cot Amount trong {os.path.basename(filepath)}")
        if not col_ten:
            col_ten = df.columns[1] if len(df.columns) > 1 else df.columns[0]
        if not col_sale:
            col_sale = df.columns[-1]

        df[col_amount] = pd.to_numeric(df[col_amount], errors="coerce").fillna(0)
        df = df[df[col_ten].notna()].copy()
        return df, col_ten, col_amount, col_sale, col_indon, col_inv, col_ngay, col_bit

    else:
        # File don le - trich xuat 1 dong
        dong = doc_file_don_le_doanhthu(filepath, ten_sale)
        df = pd.DataFrame([dong])
        col_ten = "Tên khách "
        col_amount = "Amount"
        col_sale = "Sale"
        col_indon = "In đơn"
        col_inv = "INV Date"
        col_ngay = "Ngày gửi đơn"
        col_bit = "Số Bit "
        df[col_amount] = pd.to_numeric(df[col_amount], errors="coerce").fillna(0)
        return df, col_ten, col_amount, col_sale, col_indon, col_inv, col_ngay, col_bit


def viet_sheet_quan_ly(wb, df_all, col_ten, col_amount, col_sale, col_indon, col_inv, col_ngay, col_bit):
    ws = wb.create_sheet("Quan ly")
    headers = ["STT", "Ten khach", "In don", "INV Date", "Ngay gui don", "Amount", "So Bit", "Sale"]
    ws.append(headers)
    for cell in ws[1]:
        style_header(cell)

    cols_order = [col_ten, col_indon, col_inv, col_ngay, col_amount, col_bit, col_sale]

    for i, (_, row) in enumerate(df_all.iterrows(), start=1):
        ws.cell(i + 1, 1, i).font = Font(name="Arial", size=10)
        for j, col in enumerate(cols_order, start=2):
            if col and col in df_all.columns:
                val = row[col]
                if pd.isna(val):
                    val = None
                c = ws.cell(i + 1, j, val)
                c.font = Font(name="Arial", size=10)
                if j == 6 and isinstance(val, (int, float)):
                    c.number_format = "#,##0"

    r = len(df_all) + 2
    ws.cell(r, 1, "TONG")
    ws.cell(r, 6, f"=SUM(F2:F{r-1})")
    ws.cell(r, 6).number_format = "#,##0"
    ws.cell(r, 7, f"=SUM(G2:G{r-1})")
    for col in range(1, 9):
        style_total(ws.cell(r, col))

    widths = [6, 22, 12, 12, 12, 12, 10, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w


def viet_sheet_cong_no(wb, df_all, col_ten, col_amount):
    ws = wb.create_sheet("Cong no")
    ws.append(["STT", "Ten khach hang", "Tong tien"])
    for cell in ws[1]:
        style_header(cell)

    khach = df_all.groupby(col_ten)[col_amount].sum().reset_index()
    khach.columns = ["Ten", "Tong"]
    khach = khach.sort_values("Tong", ascending=False).reset_index(drop=True)

    for i, row in khach.iterrows():
        ws.cell(i + 2, 1, i + 1).font = Font(name="Arial", size=10)
        ws.cell(i + 2, 2, row["Ten"]).font = Font(name="Arial", size=10)
        c = ws.cell(i + 2, 3, row["Tong"])
        c.font = Font(name="Arial", size=10)
        c.number_format = "#,##0"

    r = len(khach) + 2
    ws.cell(r, 1, "TONG")
    ws.cell(r, 3, f"=SUM(C2:C{r-1})")
    ws.cell(r, 3).number_format = "#,##0"
    for col in range(1, 4):
        style_total(ws.cell(r, col))

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 15


def viet_sheet_dthu(wb, df_all, col_sale, col_amount):
    ws = wb.create_sheet("DThu")
    ws.append(["Sale", "Tong doanh thu"])
    for cell in ws[1]:
        style_header(cell)

    sale = df_all.groupby(col_sale)[col_amount].sum().reset_index()
    sale.columns = ["Sale", "Tong"]
    sale = sale.sort_values("Tong", ascending=False).reset_index(drop=True)

    for i, row in sale.iterrows():
        ws.cell(i + 2, 1, row["Sale"]).font = Font(name="Arial", size=10)
        c = ws.cell(i + 2, 2, row["Tong"])
        c.font = Font(name="Arial", size=10)
        c.number_format = "#,##0"

    r = len(sale) + 2
    ws.cell(r, 1, "TONG")
    ws.cell(r, 2, f"=SUM(B2:B{r-1})")
    ws.cell(r, 2).number_format = "#,##0"
    for col in range(1, 3):
        style_total(ws.cell(r, col))

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18


def tong_hop_doanh_thu(input_files, output_file, thang=None):
    all_dfs = []
    col_ten = col_amount = col_sale = col_indon = col_inv = col_ngay = col_bit = None

    for f in input_files:
        try:
            df, ct, ca, cs, ci, cinv, cn, cb = doc_file_doanhthu(f)
            if not col_ten and ct:
                col_ten = ct
            if not col_amount and ca:
                col_amount = ca
            if not col_sale and cs:
                col_sale = cs
            col_indon = ci or col_indon
            col_inv = cinv or col_inv
            col_ngay = cn or col_ngay
            col_bit = cb or col_bit
            all_dfs.append(df)
        except Exception as e:
            print(f"Bo qua {os.path.basename(f)}: {e}")

    if not all_dfs:
        raise ValueError("Khong co file hop le nao de tong hop")

    df_all = pd.concat(all_dfs, ignore_index=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    viet_sheet_quan_ly(wb, df_all, col_ten, col_amount, col_sale, col_indon, col_inv, col_ngay, col_bit)
    if col_ten and col_amount:
        viet_sheet_cong_no(wb, df_all, col_ten, col_amount)
    if col_sale and col_amount:
        viet_sheet_dthu(wb, df_all, col_sale, col_amount)

    wb.save(output_file)
    print(f"Da tao: {output_file}")
    return output_file


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Cach dung: python tao_bao_cao_doanhthu.py file1.xlsx [file2.xlsx ...] output.xlsx")
        sys.exit(1)
    tong_hop_doanh_thu(sys.argv[1:-1], sys.argv[-1])


def them_vao_doanh_thu(dong_moi, master_file):
    """Them 1 dong vao master file doanh thu."""
    import pandas as pd
    cols = ["Tên khách ", "In đơn", "INV Date", "Ngày gửi đơn", "Amount", "Số Bit ", "Sale"]
    if os.path.exists(master_file):
        try:
            df_cu = pd.read_excel(master_file, sheet_name="Quản lý")
        except Exception:
            df_cu = pd.DataFrame(columns=cols)
    else:
        df_cu = pd.DataFrame(columns=cols)
    df_moi = pd.DataFrame([dong_moi])
    df_all = pd.concat([df_cu, df_moi], ignore_index=True)
    with pd.ExcelWriter(master_file, engine="openpyxl") as writer:
        df_all.to_excel(writer, sheet_name="Quản lý", index=False)
    return len(df_all)
