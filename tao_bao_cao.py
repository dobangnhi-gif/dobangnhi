"""
Script tổng hợp báo cáo sản lượng từ file Excel thô.
Cách dùng: python tao_bao_cao.py <input_file.xlsx> [output_file.xlsx]
"""

import sys
import os
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def doc_du_lieu(filepath):
    df = pd.read_excel(filepath, sheet_name="Tổng hợp sản lượng", header=0)
    df.columns = df.columns.str.strip()
    # Bỏ các hàng không có Product
    df = df[df["Product"].notna()].copy()
    df["Qty"] = pd.to_numeric(df["Qty"], errors="coerce").fillna(0)
    return df


def style_header(cell, bg_color="1F4E79", font_color="FFFFFF", bold=True):
    cell.font = Font(name="Arial", bold=bold, color=font_color, size=10)
    cell.fill = PatternFill("solid", start_color=bg_color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_total(cell):
    cell.font = Font(name="Arial", bold=True, size=10)
    cell.fill = PatternFill("solid", start_color="D6E4F0")


def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def viet_sheet_product(wb, df):
    ws = wb.create_sheet("Product")
    pivot = df.groupby("Product")["Qty"].sum().reset_index()
    pivot.columns = ["Loại hàng", "Số lượng"]
    pivot = pivot.sort_values("Số lượng", ascending=False)

    ws.append(["Loại hàng", "Số lượng", "Thực tế (kg)"])
    for cell in ws[1]:
        style_header(cell)

    # Hệ số quy đổi
    he_so = {"wigs 200": 3, "wigs 250": 3.5, "wigs 300": 4, "wigs 350": 4.5, "wigs 400": 5}

    for i, row in enumerate(pivot.itertuples(index=False), start=2):
        ws.cell(i, 1, row[0])
        ws.cell(i, 2, row[1])
        hs = he_so.get(str(row[0]).lower(), "")
        if hs:
            ws.cell(i, 3, row[1] * hs)
        ws.cell(i, 1).font = Font(name="Arial", size=10)
        ws.cell(i, 2).font = Font(name="Arial", size=10)

    # Tổng
    r = len(pivot) + 2
    ws.cell(r, 1, "TỔNG")
    ws.cell(r, 2, f"=SUM(B2:B{r-1})")
    style_total(ws.cell(r, 1))
    style_total(ws.cell(r, 2))

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 15


def viet_sheet_quality(wb, df):
    ws = wb.create_sheet("Quality")
    pivot = df.pivot_table(index="Product", columns="Quality", values="Qty", aggfunc="sum", fill_value=0)
    pivot["TỔNG"] = pivot.sum(axis=1)
    pivot = pivot.reset_index()

    headers = ["Loại hàng"] + [str(c) for c in pivot.columns[1:]]
    ws.append(headers)
    for cell in ws[1]:
        style_header(cell)

    for i, row in enumerate(pivot.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, val)
            ws.cell(i, j).font = Font(name="Arial", size=10)

    ws.column_dimensions["A"].width = 15
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10


def viet_sheet_length(wb, df):
    ws = wb.create_sheet("Length")
    pivot = df.pivot_table(index="Product", columns="Lenght", values="Qty", aggfunc="sum", fill_value=0)

    # Sắp xếp chiều dài
    def sort_key(x):
        try:
            return float(str(x).replace('"', '').replace("'", "").strip())
        except:
            return 9999

    sorted_cols = sorted(pivot.columns, key=sort_key)
    pivot = pivot[sorted_cols]
    pivot["TỔNG"] = pivot.sum(axis=1)
    pivot = pivot.reset_index()

    headers = ["Loại hàng"] + [str(c) for c in pivot.columns[1:]]
    ws.append(headers)
    for cell in ws[1]:
        style_header(cell)

    for i, row in enumerate(pivot.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, val)
            ws.cell(i, j).font = Font(name="Arial", size=10)

    ws.column_dimensions["A"].width = 15
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 8


def viet_sheet_color(wb, df):
    ws = wb.create_sheet("Color")

    def classify_color(c):
        if str(c).lower() in ["natural color", "natural colour"]:
            return "No Color"
        return "Color"

    df = df.copy()
    df["Loại màu"] = df["Color"].apply(classify_color)
    pivot = df.pivot_table(index="Product", columns="Loại màu", values="Qty", aggfunc="sum", fill_value=0)
    pivot["TỔNG"] = pivot.sum(axis=1)
    pivot = pivot.reset_index()

    headers = ["Loại hàng"] + [str(c) for c in pivot.columns[1:]]
    ws.append(headers)
    for cell in ws[1]:
        style_header(cell)

    for i, row in enumerate(pivot.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            ws.cell(i, j, val)
            ws.cell(i, j).font = Font(name="Arial", size=10)

    ws.column_dimensions["A"].width = 15
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12


def viet_sheet_tong_hop(wb, df, ten_bao_cao):
    ws = wb.create_sheet("Tổng hợp", 0)
    ws.sheet_view.showGridLines = False

    # Tiêu đề
    ws.merge_cells("A1:F1")
    ws["A1"] = ten_bao_cao
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"Xuất báo cáo: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="808080")

    # Tổng quan
    ws["A4"] = "TỔNG QUAN"
    ws["A4"].font = Font(name="Arial", bold=True, size=11)

    tong_qty = df["Qty"].sum()
    so_don = df["Code 1 "].nunique() if "Code 1 " in df.columns else df.index.nunique()

    data = [
        ("Tổng số lượng (sợi/cái)", tong_qty),
        ("Tổng số đơn hàng", so_don),
        ("Số loại sản phẩm", df["Product"].nunique()),
    ]

    for i, (label, val) in enumerate(data, start=5):
        ws.cell(i, 1, label).font = Font(name="Arial", size=10)
        ws.cell(i, 2, val).font = Font(name="Arial", bold=True, size=11)

    # Thống kê theo nhân viên bán hàng
    if "Sale " in df.columns:
        ws["A9"] = "THEO NHÂN VIÊN BÁN HÀNG"
        ws["A9"].font = Font(name="Arial", bold=True, size=11)
        style_header(ws["A10"])
        style_header(ws["B10"])
        ws["A10"] = "Tên"
        ws["B10"] = "Số lượng"
        sale_sum = df.groupby("Sale ")["Qty"].sum().sort_values(ascending=False)
        for i, (name, qty) in enumerate(sale_sum.items(), start=11):
            ws.cell(i, 1, name).font = Font(name="Arial", size=10)
            ws.cell(i, 2, qty).font = Font(name="Arial", size=10)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15


def tao_bao_cao(input_file, output_file=None):
    if not output_file:
        base = os.path.splitext(input_file)[0]
        output_file = f"{base}_BAOCAO_{datetime.now().strftime('%d%m%Y')}.xlsx"

    df = doc_du_lieu(input_file)

    # Lấy khoảng ngày từ dữ liệu
    if "Ngày đưa đơn" in df.columns:
        dates = pd.to_datetime(df["Ngày đưa đơn"], errors="coerce").dropna()
        if len(dates):
            ten_bao_cao = f"BÁO CÁO SẢN LƯỢNG {dates.min().strftime('%d/%m')} - {dates.max().strftime('%d/%m/%Y')}"
        else:
            ten_bao_cao = "BÁO CÁO SẢN LƯỢNG"
    else:
        ten_bao_cao = "BÁO CÁO SẢN LƯỢNG"

    wb = openpyxl.Workbook()
    # Xóa sheet mặc định
    wb.remove(wb.active)

    viet_sheet_tong_hop(wb, df, ten_bao_cao)
    viet_sheet_product(wb, df)
    viet_sheet_quality(wb, df)
    viet_sheet_length(wb, df)
    viet_sheet_color(wb, df)

    wb.save(output_file)
    print(f"✅ Đã tạo báo cáo: {output_file}")
    return output_file


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Cách dùng: python tao_bao_cao.py <input.xlsx> [output.xlsx]")
        sys.exit(1)
    tao_bao_cao(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
