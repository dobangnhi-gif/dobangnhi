from flask import Flask, request, send_file, jsonify, render_template_string
import os, sys
from datetime import datetime, timedelta

sys.path.insert(0, os.path.dirname(__file__))
from doc_file_don_le import doc_file_don_le, them_vao_tong_hop
from tao_bao_cao import tao_bao_cao
from tao_bao_cao_doanhthu import doc_file_don_le_doanhthu, them_vao_doanh_thu, viet_sheet_quan_ly, viet_sheet_cong_no, viet_sheet_dthu
from github_storage import upload_file as gh_upload, download_file as gh_download
import pandas as pd
import openpyxl

app = Flask(__name__)
UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

MASTER_FILENAME = "tong_hop_san_luong.xlsx"
MASTER_FILE = os.path.join(UPLOAD_DIR, MASTER_FILENAME)
DOANHTHU_FILENAME = "master_doanh_thu.xlsx"
DOANHTHU_FILE = os.path.join(UPLOAD_DIR, DOANHTHU_FILENAME)

HTML = """
<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>IndonBot - Quản lý đơn hàng</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: Arial, sans-serif; background: #f0f2f5; min-height: 100vh; display: flex; align-items: center; justify-content: center; padding: 20px; }
  .card { background: white; border-radius: 16px; padding: 30px; max-width: 520px; width: 100%; box-shadow: 0 4px 20px rgba(0,0,0,0.1); }
  h1 { color: #1a73e8; font-size: 22px; margin-bottom: 6px; }
  p.sub { color: #666; margin-bottom: 20px; font-size: 13px; }
  .tab-bar { display: flex; gap: 8px; margin-bottom: 20px; flex-wrap: wrap; }
  .tab { padding: 7px 14px; border-radius: 20px; cursor: pointer; font-size: 13px; border: 2px solid #1a73e8; color: #1a73e8; background: white; font-weight: bold; }
  .tab.active { background: #1a73e8; color: white; }
  .section { display: none; }
  .section.active { display: block; }
  .upload-area { border: 2px dashed #1a73e8; border-radius: 10px; padding: 25px; text-align: center; cursor: pointer; }
  .upload-area:hover { background: #f0f7ff; }
  .upload-area input { display: none; }
  .upload-area .icon { font-size: 36px; margin-bottom: 8px; }
  .upload-area label { cursor: pointer; color: #1a73e8; font-weight: bold; font-size: 14px; }
  #file-name { margin-top: 8px; color: #333; font-size: 12px; }
  .btn { display: block; width: 100%; padding: 12px; background: #1a73e8; color: white; border: none; border-radius: 8px; font-size: 15px; font-weight: bold; cursor: pointer; margin-top: 15px; }
  .btn:hover { background: #1557b0; }
  .btn:disabled { background: #aaa; cursor: not-allowed; }
  .btn.green { background: #43a047; }
  .btn.green:hover { background: #2e7d32; }
  .result { margin-top: 15px; padding: 12px; border-radius: 8px; display: none; font-size: 13px; }
  .result.success { background: #e8f5e9; color: #2e7d32; }
  .result.error { background: #ffebee; color: #c62828; }
  .download-btn { display: inline-block; padding: 8px 16px; background: #43a047; color: white; border-radius: 6px; text-decoration: none; font-weight: bold; margin-top: 8px; font-size: 13px; }
  label.lbl { color: #555; font-size: 12px; font-weight: bold; display: block; margin-top: 12px; }
  select, input[type=text] { width: 100%; padding: 9px; border: 2px solid #ddd; border-radius: 7px; font-size: 13px; margin-top: 5px; }
  .row2 { display: flex; gap: 10px; }
  .row2 > div { flex: 1; }
</style>
</head>
<body>
<div class="card">
  <h1>🗂️ IndonBot</h1>
  <p class="sub">Hệ thống tổng hợp đơn hàng</p>

  <div class="tab-bar">
    <button class="tab active" onclick="showTab('upload', this)">📤 Gửi đơn</button>
    <button class="tab" onclick="showTab('sanluong', this)">📦 Sản lượng</button>
  </div>

  <!-- Tab gửi đơn -->
  <div id="tab-upload" class="section active">
    <div class="upload-area" onclick="document.getElementById('file-input').click()">
      <div class="icon">📎</div>
      <label>Nhấn để chọn file Excel (.xlsx)</label>
      <input type="file" id="file-input" accept=".xlsx,.xls" onchange="onFileSelect(this)">
      <div id="file-name">Chưa chọn file</div>
    </div>
    <label class="lbl">Tên của bạn (Sale):</label>
    <input type="text" id="sale-name" placeholder="Ví dụ: A Thắng">
    <button class="btn" id="upload-btn" onclick="uploadFile()" disabled>📤 Gửi đơn</button>
    <div class="result" id="upload-result"></div>
  </div>

  <!-- Tab sản lượng -->
  <div id="tab-sanluong" class="section">
    <label class="lbl">Chọn kỳ:</label>
    <select id="sl-type" onchange="toggleDates('sl')">
      <option value="today">Hôm nay</option>
      <option value="week">Tuần này</option>
      <option value="custom">Khoảng ngày tùy chọn</option>
    </select>
    <div id="sl-dates" style="display:none">
      <div class="row2">
        <div><label class="lbl">Từ ngày (DD/MM):</label><input type="text" id="sl-from" placeholder="01/06"></div>
        <div><label class="lbl">Đến ngày (DD/MM):</label><input type="text" id="sl-to" placeholder="07/06"></div>
      </div>
    </div>
    <button class="btn" onclick="getReport('sanluong')">📊 Tạo báo cáo sản lượng</button>
    <div class="result" id="sl-result"></div>
  </div>


</div>

<script>
function showTab(name, btn) {
  document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
  document.querySelectorAll('.section').forEach(s => s.classList.remove('active'));
  document.getElementById('tab-' + name).classList.add('active');
  btn.classList.add('active');
}

function toggleDates(prefix) {
  const v = document.getElementById(prefix + '-type').value;
  document.getElementById(prefix + '-dates').style.display = v === 'custom' ? 'block' : 'none';
}

function onFileSelect(input) {
  const f = input.files[0];
  document.getElementById('file-name').textContent = f ? f.name : 'Chưa chọn file';
  document.getElementById('upload-btn').disabled = !f;
}

async function uploadFile() {
  const file = document.getElementById('file-input').files[0];
  const sale = document.getElementById('sale-name').value || 'Không rõ';
  const btn = document.getElementById('upload-btn');
  const result = document.getElementById('upload-result');
  btn.disabled = true; btn.textContent = '⏳ Đang xử lý...';
  result.style.display = 'none';
  const formData = new FormData();
  formData.append('file', file);
  formData.append('sale', sale);
  try {
    const res = await fetch('/upload', { method: 'POST', body: formData });
    const data = await res.json();
    result.style.display = 'block';
    if (data.success) {
      result.className = 'result success';
      result.innerHTML = '✅ ' + data.message + '<br><a class="download-btn" href="' + data.download_url + '" download>⬇️ Tải báo cáo</a>';
    } else {
      result.className = 'result error';
      result.textContent = '❌ ' + data.error;
    }
  } catch(e) {
    result.style.display = 'block'; result.className = 'result error';
    result.textContent = '❌ Lỗi kết nối';
  }
  btn.disabled = false; btn.textContent = '📤 Gửi đơn';
}

async function getReport(type) {
  const resultId = type === 'doanhthu' ? 'dt-result' : 'sl-result';
  const result = document.getElementById(resultId);
  result.style.display = 'block'; result.className = 'result success';
  result.textContent = '⏳ Đang tạo báo cáo...';
  let params = 'type=' + type;
  if (type === 'sanluong') {
    const slType = document.getElementById('sl-type').value;
    params += '&period=' + slType;
    if (slType === 'custom') {
      params += '&from=' + document.getElementById('sl-from').value;
      params += '&to=' + document.getElementById('sl-to').value;
    }
  } else {
    params += '&month=' + document.getElementById('dt-month').value;
    params += '&year=' + document.getElementById('dt-year').value;
  }
  try {
    const res = await fetch('/report?' + params);
    const data = await res.json();
    if (data.success) {
      result.innerHTML = '✅ ' + data.message + '<br><a class="download-btn" href="' + data.download_url + '" download>⬇️ Tải báo cáo</a>';
    } else {
      result.className = 'result error'; result.textContent = '❌ ' + data.error;
    }
  } catch(e) {
    result.className = 'result error'; result.textContent = '❌ Lỗi kết nối';
  }
}
</script>
</body>
</html>
"""


@app.route('/')
def index():
    return render_template_string(HTML)


@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({"success": False, "error": "Không có file"})
    f = request.files['file']
    sale = request.form.get('sale', 'Không rõ')
    if not f.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"success": False, "error": "Chỉ chấp nhận file Excel (.xlsx)"})
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    save_path = os.path.join(UPLOAD_DIR, f"{ts}_{f.filename}")
    f.save(save_path)
    try:
        if not os.path.exists(MASTER_FILE):
            gh_download(MASTER_FILENAME, MASTER_FILE)
        du_lieu, ten_kh = doc_file_don_le(save_path, sale)
        ngay_gui = datetime.now()
        for d in du_lieu:
            d["Ngày đưa đơn"] = ngay_gui
        tong = them_vao_tong_hop(du_lieu, MASTER_FILE)
        gh_upload(MASTER_FILE, MASTER_FILENAME)

        # Tong hop doanh thu
        try:
            if not os.path.exists(DOANHTHU_FILE):
                gh_download(DOANHTHU_FILENAME, DOANHTHU_FILE)
            dong_dt = doc_file_don_le_doanhthu(save_path, sale)
            dong_dt["Ngày gửi đơn"] = ngay_gui
            them_vao_doanh_thu(dong_dt, DOANHTHU_FILE)
            gh_upload(DOANHTHU_FILE, DOANHTHU_FILENAME)
        except Exception as e:
            pass  # Doanh thu khong bat buoc

        # Tao bao cao cho file nay
        tmp = save_path + "_tmp.xlsx"
        df = pd.DataFrame(du_lieu)
        df.to_excel(tmp, sheet_name="Tổng hợp sản lượng", index=False)
        out = save_path.replace(".xlsx", "_baocao.xlsx")
        tao_bao_cao(tmp, out)
        wb = openpyxl.load_workbook(out)
        ws_raw = wb.create_sheet("Chi tiet", 0)
        ws_raw.append(list(df.columns))
        for _, row in df.iterrows():
            ws_raw.append([row[c] for c in df.columns])
        wb.save(out)
        return jsonify({
            "success": True,
            "message": f"Đã lưu {len(du_lieu)} sản phẩm của {ten_kh}. Tổng: {tong} dòng",
            "download_url": f"/download/{os.path.basename(out)}"
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route('/report')
def report():
    report_type = request.args.get('type', 'sanluong')
    now = datetime.now()

    if report_type == 'doanhthu':
        # Báo cáo doanh thu theo tháng
        try:
            thang = int(request.args.get('month', now.month))
            nam = int(request.args.get('year', now.year))
        except Exception:
            return jsonify({"success": False, "error": "Tháng/năm không hợp lệ"})

        if not os.path.exists(DOANHTHU_FILE):
            result = gh_download(DOANHTHU_FILENAME, DOANHTHU_FILE)
            if not result:
                return jsonify({"success": False, "error": "Chưa có dữ liệu doanh thu."})
        try:
            df = pd.read_excel(DOANHTHU_FILE, sheet_name="Quản lý")
            col_ngay = next((c for c in df.columns if "ngay" in str(c).lower() and "gui" in str(c).lower()), None)
            if col_ngay:
                df[col_ngay] = pd.to_datetime(df[col_ngay], errors="coerce")
                tu_ngay = datetime(nam, thang, 1)
                den_ngay = datetime(nam, thang + 1, 1) - timedelta(seconds=1) if thang < 12 else datetime(nam + 1, 1, 1) - timedelta(seconds=1)
                df_filter = df[(df[col_ngay] >= tu_ngay) & (df[col_ngay] <= den_ngay)]
            else:
                df_filter = df

            if df_filter.empty:
                return jsonify({"success": False, "error": f"Không có dữ liệu doanh thu tháng {thang}/{nam}"})

            out = os.path.join(UPLOAD_DIR, f"DoanhThu_T{thang:02d}_{nam}.xlsx")
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            col_ten = "Tên khách "
            col_amount = "Amount"
            col_sale = "Sale"
            col_indon = "In đơn"
            col_inv = "INV Date"
            col_ngay2 = "Ngày gửi đơn"
            col_bit = "Số Bit "
            viet_sheet_quan_ly(wb, df_filter, col_ten, col_amount, col_sale, col_indon, col_inv, col_ngay2, col_bit)
            if col_ten in df_filter.columns and col_amount in df_filter.columns:
                viet_sheet_cong_no(wb, df_filter, col_ten, col_amount)
            if col_sale in df_filter.columns and col_amount in df_filter.columns:
                viet_sheet_dthu(wb, df_filter, col_sale, col_amount)
            wb.save(out)

            tong_amount = df_filter[col_amount].sum() if col_amount in df_filter.columns else 0
            return jsonify({
                "success": True,
                "message": f"Doanh thu T{thang}/{nam}: {len(df_filter)} đơn | Tổng: {tong_amount:,.0f}",
                "download_url": f"/download/{os.path.basename(out)}"
            })
        except Exception as e:
            return jsonify({"success": False, "error": str(e)})

    else:
        # Báo cáo sản lượng
        if not os.path.exists(MASTER_FILE):
            result = gh_download(MASTER_FILENAME, MASTER_FILE)
            if not result:
                return jsonify({"success": False, "error": "Chưa có dữ liệu sản lượng."})
        try:
            try:
                df = pd.read_excel(MASTER_FILE, sheet_name="Tổng hợp sản lượng")
            except Exception:
                df = pd.read_excel(MASTER_FILE, sheet_name=0)
            df["Ngày đưa đơn"] = pd.to_datetime(df["Ngày đưa đơn"], errors="coerce")

            period = request.args.get('period', 'today')
            if period == 'today':
                df_filter = df[df["Ngày đưa đơn"].dt.date == now.date()]
                ten_ky = f"hom-nay-{now.strftime('%d%m%Y')}"
            elif period == 'week':
                monday = now.date() - timedelta(days=now.weekday())
                df_filter = df[df["Ngày đưa đơn"].dt.date >= monday]
                ten_ky = f"tuan-{monday.strftime('%d%m')}"
            else:
                from_str = request.args.get('from', '')
                to_str = request.args.get('to', '')
                try:
                    d1 = datetime.strptime(f"{from_str}/{now.year}", "%d/%m/%Y").date()
                    d2 = datetime.strptime(f"{to_str}/{now.year}", "%d/%m/%Y").date()
                    df_filter = df[(df["Ngày đưa đơn"].dt.date >= d1) & (df["Ngày đưa đơn"].dt.date <= d2)]
                    ten_ky = f"{from_str.replace('/','-')}-{to_str.replace('/','-')}"
                except Exception:
                    return jsonify({"success": False, "error": "Định dạng ngày không đúng (DD/MM)"})

            if df_filter.empty:
                return jsonify({"success": False, "error": "Không có dữ liệu trong khoảng thời gian này"})

            tmp = os.path.join(UPLOAD_DIR, "temp_report.xlsx")
            df_filter.to_excel(tmp, sheet_name="Tổng hợp sản lượng", index=False)
            out = os.path.join(UPLOAD_DIR, f"BaoCao_{ten_ky}.xlsx")
            tao_bao_cao(tmp, out)
            wb = openpyxl.load_workbook(out)
            ws_raw = wb.create_sheet("Chi tiet", 0)
            ws_raw.append(list(df_filter.columns))
            for _, row in df_filter.iterrows():
                ws_raw.append([row[c] for c in df_filter.columns])
            wb.save(out)
            return jsonify({
                "success": True,
                "message": f"Báo cáo sản lượng: {len(df_filter)} dòng",
                "download_url": f"/download/{os.path.basename(out)}"
            })
        except Exception as e:
            return jsonify({"success": False, "error": str(e)})


@app.route('/download/<filename>')
def download(filename):
    path = os.path.join(UPLOAD_DIR, filename)
    if os.path.exists(path):
        return send_file(path, as_attachment=True)
    return "File không tìm thấy", 404


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
